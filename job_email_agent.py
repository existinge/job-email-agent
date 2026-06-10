#!/usr/bin/env python3
"""
Job Hunt Outlook Email Agent + OpenRouter AI

Modes:
  review           Read recent mail, classify with OpenRouter when configured, and export Excel. No mailbox changes.
  apply            Do review + move high-confidence Rejections and Job Board Spam into Outlook folders.
  draft-approved   Read an edited Excel report and create Outlook drafts for rows marked Approved To Send = YES.
  send-approved    Read an edited Excel report and send only rows marked Approved To Send = YES.
  mark-report-processed Mark rows from an existing report as already reviewed/processed.
  processed-status Show how many emails are currently logged as processed.
  reset-processed  Clear the processed-message log.
  init             Create config.yaml and .env.template.

The script uses Microsoft Graph delegated permissions through MSAL device-code login.
OpenRouter is optional. If no OpenRouter key is present, it falls back to local rules.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import hashlib
import html
import json
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlencode

import msal
import requests
import yaml
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

GRAPH_ROOT = "https://graph.microsoft.com/v1.0"
OPENROUTER_ROOT = "https://openrouter.ai/api/v1/chat/completions"
TOKEN_CACHE_FILE = ".msal_token_cache.json"
AI_CACHE_FILE = ".openrouter_classification_cache.json"
PROCESSED_LOG_FILE = "processed_messages.json"
SENT_LOG_FILE = "sent_followup_log.json"

# MSAL device-code flow treats openid/profile/offline_access as reserved scopes.
# Keep only Graph API scopes here.
DEFAULT_SCOPES = ["User.Read", "Mail.ReadWrite", "Mail.Send"]
RESERVED_MSAL_SCOPES = {"offline_access", "openid", "profile"}

ALLOWED_CATEGORIES = {
    "Needs Action",
    "Follow Up Queue",
    "Application Receipt",
    "Rejection",
    "Job Board Spam",
    "Important Job Email",
    "Ignore / Non-Job",
}

DEFAULT_CONFIG = {
    "user": {
        "your_name": "Your Name",
        "signature": "Best,\nYour Name",
        "follow_up_tone": "warm, concise, professional, not pushy",
        "job_search_focus": "remote support, customer support, helpdesk, intake, chat/email roles",
    },
    "graph": {
        "scopes": DEFAULT_SCOPES,
        "tenant": "consumers",  # Use consumers for personal Outlook/Hotmail. Use common for mixed account types.
        "page_size": 50,
    },
    "scan": {
        "days_back": 60,
        "max_messages": 250,
        "source": "inbox",  # inbox or all
        "include_body_text": True,
        "include_non_job_related_in_all_sheet": True,
    },
    "processed": {
        "enabled": True,
        "file": PROCESSED_LOG_FILE,
        "skip_previously_processed": True,
        "mark_as_processed_after_report": True,
    },
    "ai": {
        "enabled": True,
        "provider": "openrouter",
        "model_env": "OPENROUTER_MODEL",
        "api_key_env": "OPENROUTER_API_KEY",
        "default_model": "openrouter/auto",
        "temperature": 0.15,
        "max_body_chars": 4500,
        "timeout_seconds": 90,
        "retries": 3,
        "use_cache": True,
        "send_email_content_to_openrouter": True,
        "fallback_to_rules_if_ai_fails": True,
        "only_ai_for_job_likely_messages": False,
        "min_ai_confidence_to_override_rules": 0.55,
    },
    "folders": {
        "root": "Job Hunt",
        "rejections": "Rejections",
        "job_board_spam": "Job Board Spam",
        "follow_up": "Follow Up",
        "important": "Important",
    },
    "automation": {
        "auto_move_rejections": True,
        "auto_move_job_board_spam": True,
        "auto_move_follow_up": False,
        "min_confidence_to_move": 0.72,
        "never_delete_automatically": True,
    },
    "follow_up": {
        "application_receipt_wait_days": 7,
        "recruiter_wait_days": 3,
        "default_wait_days": 5,
        "send_only_if_due": True,
        "require_approval_yes": True,
    },
    "classification": {
        "job_board_domains": [
            "indeed.com",
            "ziprecruiter.com",
            "linkedin.com",
            "monster.com",
            "careerbuilder.com",
            "glassdoor.com",
            "dice.com",
            "hired.com",
            "talent.com",
            "simplyhired.com",
            "jooble.org",
            "lensa.com",
        ],
        "spam_sender_phrases": [
            "job alert",
            "recommended jobs",
            "new jobs",
            "jobs you may like",
            "hiring now",
            "daily jobs",
            "weekly jobs",
            "similar jobs",
            "job recommendations",
            "remote jobs for you",
        ],
        "rejection_phrases": [
            "unfortunately",
            "not selected",
            "not be moving forward",
            "not moving forward",
            "decided to move forward with other candidates",
            "pursue other candidates",
            "other candidates",
            "will not be proceeding",
            "no longer under consideration",
            "we are unable to offer",
            "we will not be advancing",
            "position has been filled",
            "after careful consideration",
            "regret to inform",
            "not the right fit",
            "not proceed",
        ],
        "important_action_phrases": [
            "select a time",
            "schedule an interview",
            "interview invitation",
            "invite you to interview",
            "availability for an interview",
            "calendar invite",
            "calendly",
            "book a time",
            "phone screen",
            "action required",
            "please complete",
            "complete this assessment",
            "complete the assessment",
            "take the assessment",
            "reply with your availability",
            "are you available",
            "we'd like to speak",
            "we would like to speak",
            "congratulations",
            "offer",
        ],
        "application_receipt_phrases": [
            "thank you for applying",
            "thanks for applying",
            "we received your application",
            "application received",
            "your application has been received",
            "application submitted",
            "we have received your resume",
            "thank you for your interest",
            "your application is being reviewed",
            "we are reviewing your application",
            "our team will review your application",
            "if your qualifications match",
            "we will be in touch",
            "we’ll be in touch",
            "we'll be in touch",
            "thanks for your interest in",
            "application for the",
            "received your resume",
        ],
        "job_relevant_phrases": [
            "application",
            "resume",
            "interview",
            "recruiter",
            "hiring",
            "talent acquisition",
            "candidate",
            "position",
            "role",
            "job",
            "employment",
            "workday",
            "greenhouse",
            "lever",
            "bamboohr",
            "icims",
            "ashby",
            "smartrecruiters",
            "workable",
        ],
        "ignore_domains": [],
        "trusted_human_domains": [],
    },
}


@dataclasses.dataclass
class EmailRecord:
    id: str
    conversation_id: str
    subject: str
    received: str
    sender_name: str
    sender_email: str
    preview: str
    body_text: str
    web_link: str
    internet_message_id: str = ""
    importance: str = "normal"
    has_attachments: bool = False


@dataclasses.dataclass
class Classification:
    category: str
    confidence: float
    reason: str
    suggested_action: str
    follow_up_date: str
    approved_to_send: str
    follow_up_subject: str
    follow_up_body: str
    target_folder_key: str
    move_eligible: bool
    company_guess: str
    ai_used: bool = False
    ai_model: str = ""
    ai_error: str = ""


class GraphClient:
    def __init__(self, client_id: str, tenant: str, scopes: List[str], cache_path: Path):
        self.client_id = client_id
        self.tenant = tenant
        self.scopes = scopes
        self.cache_path = cache_path
        self.token_cache = msal.SerializableTokenCache()
        if cache_path.exists():
            self.token_cache.deserialize(cache_path.read_text(encoding="utf-8"))
        authority = f"https://login.microsoftonline.com/{tenant}"
        self.app = msal.PublicClientApplication(
            client_id=client_id,
            authority=authority,
            token_cache=self.token_cache,
        )
        self._token: Optional[str] = None

    def save_cache(self) -> None:
        if self.token_cache.has_state_changed:
            self.cache_path.write_text(self.token_cache.serialize(), encoding="utf-8")

    def token(self) -> str:
        if self._token:
            return self._token
        result = None
        accounts = self.app.get_accounts()
        if accounts:
            result = self.app.acquire_token_silent(self.scopes, account=accounts[0])
        if not result:
            flow = self.app.initiate_device_flow(scopes=self.scopes)
            if "user_code" not in flow:
                raise RuntimeError(f"Failed to create device flow: {json.dumps(flow, indent=2)}")
            print("\nMicrosoft sign-in required:")
            print(flow["message"])
            print()
            result = self.app.acquire_token_by_device_flow(flow)
        self.save_cache()
        if "access_token" not in result:
            raise RuntimeError(f"Could not authenticate: {json.dumps(result, indent=2)}")
        self._token = result["access_token"]
        return self._token

    def request(self, method: str, url: str, **kwargs: Any) -> Any:
        if url.startswith("/"):
            url = GRAPH_ROOT + url
        headers = kwargs.pop("headers", {}) or {}
        headers["Authorization"] = f"Bearer {self.token()}"
        headers.setdefault("Content-Type", "application/json")
        for attempt in range(5):
            response = requests.request(method, url, headers=headers, timeout=60, **kwargs)
            if response.status_code in (429, 500, 502, 503, 504):
                retry_after = response.headers.get("Retry-After")
                sleep_for = int(retry_after) if retry_after and retry_after.isdigit() else 2**attempt
                time.sleep(min(sleep_for, 30))
                continue
            if response.status_code == 204:
                return None
            if response.ok:
                return response.json()
            raise RuntimeError(f"Graph API error {response.status_code} for {method} {url}:\n{response.text}")
        raise RuntimeError(f"Graph API failed after retries: {method} {url}")

    def paged_get(self, url: str, limit: Optional[int] = None, headers: Optional[dict] = None) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        next_url = url
        while next_url:
            data = self.request("GET", next_url, headers=headers or {})
            batch = data.get("value", [])
            items.extend(batch)
            if limit and len(items) >= limit:
                return items[:limit]
            next_url = data.get("@odata.nextLink")
        return items


class OpenRouterAgent:
    def __init__(self, api_key: str, model: str, config: dict, cache_path: Path):
        self.api_key = api_key
        self.model = model
        self.config = config
        self.cache_path = cache_path
        self.cache: Dict[str, Any] = {}
        if config["ai"].get("use_cache", True) and cache_path.exists():
            try:
                self.cache = json.loads(cache_path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                self.cache = {}

    def save_cache(self) -> None:
        if self.config["ai"].get("use_cache", True):
            self.cache_path.write_text(json.dumps(self.cache, indent=2, sort_keys=True), encoding="utf-8")

    def classify(self, record: EmailRecord, rule_classification: Classification) -> Classification:
        cache_key = self._cache_key(record, rule_classification)
        if self.config["ai"].get("use_cache", True) and cache_key in self.cache:
            try:
                return classification_from_ai_payload(
                    self.cache[cache_key],
                    fallback=rule_classification,
                    model=self.model,
                    cached=True,
                )
            except Exception:
                pass

        payload = self._request_payload(record, rule_classification)
        data = self._post(payload)
        text = extract_openrouter_text(data)
        ai_payload = parse_json_object(text)
        if self.config["ai"].get("use_cache", True):
            self.cache[cache_key] = ai_payload
            self.save_cache()
        return classification_from_ai_payload(ai_payload, fallback=rule_classification, model=self.model, cached=False)

    def _cache_key(self, record: EmailRecord, rule_classification: Classification) -> str:
        base = "\n".join(
            [
                self.model,
                record.id,
                record.received,
                record.sender_email,
                record.subject,
                record.preview,
                record.body_text[: int(self.config["ai"].get("max_body_chars", 4500))],
                rule_classification.category,
            ]
        )
        return hashlib.sha256(base.encode("utf-8", errors="ignore")).hexdigest()

    def _request_payload(self, record: EmailRecord, rule_classification: Classification) -> dict:
        user_cfg = self.config["user"]
        ai_cfg = self.config["ai"]
        body = compact_text(record.body_text or record.preview, int(ai_cfg.get("max_body_chars", 4500)))
        signature = user_cfg.get("signature", "Best,\nYour Name")
        today = dt.datetime.now().date().isoformat()
        system = (
            "You are a cautious job-hunt inbox triage assistant. Classify email for a job seeker, "
            "detect spam/job-board noise, identify rejections, identify real recruiter/action emails, "
            "and draft short follow-up language only when a reply would be reasonable. "
            "Return exactly one valid JSON object. Do not wrap in markdown. Do not include extra text. "
            "Never approve sending. approved_to_send must always be NO. Do not invent facts. "
            "If the message is a rejection or automated spam, leave follow_up_subject and follow_up_body empty. "
            "Use a warm, concise, professional tone."
        )
        user = {
            "today": today,
            "allowed_categories": sorted(ALLOWED_CATEGORIES),
            "category_meanings": {
                "Needs Action": "Interview invite, scheduling request, assessment, offer, or clear required reply.",
                "Follow Up Queue": "A job-related thread where a polite follow-up is due or soon due.",
                "Application Receipt": "Application received/submitted confirmation; not urgent yet unless old enough.",
                "Rejection": "The company is declining or not moving forward.",
                "Job Board Spam": "Automated job alerts, recommendations, blasts, duplicate listings, low-value boards.",
                "Important Job Email": "Useful job-related email but no immediate reply required.",
                "Ignore / Non-Job": "Not related to the job search.",
            },
            "desired_json_schema": {
                "category": "one of allowed_categories",
                "confidence": "number from 0.0 to 1.0",
                "reason": "short explanation",
                "suggested_action": "short action recommendation",
                "follow_up_date": "YYYY-MM-DD or empty string",
                "approved_to_send": "NO",
                "follow_up_subject": "subject for follow-up or empty string",
                "follow_up_body": "complete email body or empty string",
                "target_folder_key": "rejections, job_board_spam, follow_up, important, or empty string",
                "move_eligible": "true only for obvious Rejection or Job Board Spam",
                "company_guess": "company name if inferable, otherwise empty string",
            },
            "job_search_focus": user_cfg.get("job_search_focus", ""),
            "follow_up_tone": user_cfg.get("follow_up_tone", "warm, concise, professional"),
            "signature_to_use": signature,
            "local_rule_guess": dataclasses.asdict(rule_classification),
            "email": {
                "received": record.received,
                "from_name": record.sender_name,
                "from_email": record.sender_email,
                "subject": record.subject,
                "preview": record.preview,
                "body_text_truncated": body,
                "importance": record.importance,
                "has_attachments": record.has_attachments,
            },
            "follow_up_rules": self.config.get("follow_up", {}),
            "safety_rules": [
                "If unsure between human recruiter and job alert, prefer Important Job Email or Needs Action over spam.",
                "If unsure whether to send, leave approved_to_send as NO and suggest review.",
                "Do not write pushy follow-ups. Avoid over-apologizing.",
                "Never tell the user to click suspicious links. Recommend reviewing in Outlook if links are involved.",
            ],
        }
        return {
            "model": self.model,
            "messages": [
                {"role": "system", "content": system},
                {"role": "user", "content": json.dumps(user, ensure_ascii=False)},
            ],
            "temperature": float(ai_cfg.get("temperature", 0.15)),
        }

    def _post(self, payload: dict) -> dict:
        ai_cfg = self.config["ai"]
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }
        site_url = os.getenv("OPENROUTER_SITE_URL", "").strip()
        app_name = os.getenv("OPENROUTER_APP_NAME", "Job Hunt Email Agent").strip()
        if site_url:
            headers["HTTP-Referer"] = site_url
        if app_name:
            headers["X-OpenRouter-Title"] = app_name

        retries = int(ai_cfg.get("retries", 3))
        timeout = int(ai_cfg.get("timeout_seconds", 90))
        last_error = None
        for attempt in range(retries):
            try:
                response = requests.post(OPENROUTER_ROOT, headers=headers, json=payload, timeout=timeout)
                if response.status_code in (429, 500, 502, 503, 504):
                    retry_after = response.headers.get("Retry-After")
                    sleep_for = int(retry_after) if retry_after and retry_after.isdigit() else 2**attempt
                    time.sleep(min(sleep_for, 30))
                    continue
                if not response.ok:
                    raise RuntimeError(f"OpenRouter error {response.status_code}: {response.text[:1000]}")
                return response.json()
            except Exception as exc:
                last_error = exc
                if attempt < retries - 1:
                    time.sleep(min(2**attempt, 10))
        raise RuntimeError(f"OpenRouter request failed: {last_error}")


def deep_merge(base: dict, override: dict) -> dict:
    out = dict(base)
    for k, v in override.items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def load_config(path: Path) -> dict:
    config = DEFAULT_CONFIG
    if path.exists():
        loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        config = deep_merge(config, loaded)
    return config


def save_default_config(path: Path) -> None:
    if not path.exists():
        path.write_text(yaml.safe_dump(DEFAULT_CONFIG, sort_keys=False), encoding="utf-8")


def norm(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())


def strip_html(value: str) -> str:
    value = html.unescape(value or "")
    value = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", value)
    value = re.sub(r"(?s)<br\s*/?>", "\n", value)
    value = re.sub(r"(?s)</p>", "\n", value)
    value = re.sub(r"(?s)<.*?>", " ", value)
    return re.sub(r"[ \t]+", " ", value).strip()


def compact_text(value: str, limit: int) -> str:
    value = re.sub(r"\s+", " ", strip_html(value or "")).strip()
    if len(value) <= limit:
        return value
    return value[: limit - 20] + " ...[truncated]"


def sender_domain(email: str) -> str:
    email = (email or "").lower()
    return email.split("@", 1)[1] if "@" in email else ""


NO_REPLY_LOCAL_PARTS = {
    "noreply",
    "no-reply",
    "donotreply",
    "do-not-reply",
    "notification",
    "notifications",
    "automated",
    "mailer",
}

NO_REPLY_DOMAINS = {
    "ashbyhq.com",
    "hire.lever.co",
    "greenhouse-mail.io",
    "us.greenhouse-mail.io",
    "workday.com",
    "myworkday.com",
    "smartrecruiters.com",
    "bamboohr.com",
    "icims.com",
    "jobvite.com",
    "workablemail.com",
}


def is_no_reply_address(email: str) -> bool:
    email = (email or "").strip().lower()
    if "@" not in email:
        return True
    local, domain = email.split("@", 1)
    local_clean = re.sub(r"[^a-z0-9-]", "", local)
    if local_clean in NO_REPLY_LOCAL_PARTS:
        return True
    if any(token in local_clean for token in ["noreply", "no-reply", "donotreply", "do-not-reply"]):
        return True
    return any(domain == d or domain.endswith("." + d) for d in NO_REPLY_DOMAINS)


def contains_any(text: str, phrases: Iterable[str]) -> Tuple[bool, List[str]]:
    found = []
    text_l = norm(text)
    for phrase in phrases:
        p = norm(phrase)
        if p and p in text_l:
            found.append(phrase)
    return bool(found), found


def safe_date(date_str: str) -> dt.datetime:
    if not date_str:
        return dt.datetime.now(dt.timezone.utc)
    s = date_str.replace("Z", "+00:00")
    try:
        value = dt.datetime.fromisoformat(s)
        if value.tzinfo is None:
            value = value.replace(tzinfo=dt.timezone.utc)
        return value
    except ValueError:
        return dt.datetime.now(dt.timezone.utc)


def load_json_dict(path: Path) -> dict:
    if path.exists():
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
            return data if isinstance(data, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def save_json_dict(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")


def processed_log_path(project_dir: Path, config: dict) -> Path:
    filename = config.get("processed", {}).get("file", PROCESSED_LOG_FILE) or PROCESSED_LOG_FILE
    return project_dir / filename


def message_fingerprint(record: EmailRecord) -> str:
    stable_id = norm(record.internet_message_id)
    if stable_id:
        source = f"internetMessageId|{stable_id}"
    else:
        received_date = safe_date(record.received).date().isoformat() if record.received else ""
        source = "|".join(
            [
                "fallback",
                norm(record.sender_email),
                norm(record.subject),
                received_date,
                norm(record.preview)[:500],
            ]
        )
    return hashlib.sha256(source.encode("utf-8", errors="ignore")).hexdigest()


def row_processed_key(row: Dict[str, Any]) -> str:
    key = str(row.get("Processed Key") or "").strip()
    if key:
        return key
    source = "|".join(
        [
            str(row.get("Sender Email") or "").strip().lower(),
            str(row.get("Subject") or "").strip().lower(),
            str(row.get("Received") or "")[:10],
            str(row.get("Preview") or "")[:500].strip().lower(),
        ]
    )
    return hashlib.sha256(source.encode("utf-8", errors="ignore")).hexdigest()


def mark_rows_processed(path: Path, rows: List[Dict[str, Any]], run_meta: Optional[dict] = None) -> int:
    log = load_json_dict(path)
    now = dt.datetime.now(dt.timezone.utc).isoformat()
    added = 0
    for row in rows:
        key = row_processed_key(row)
        if not key:
            continue
        if key not in log:
            added += 1
        log[key] = {
            "processed_at": now,
            "message_id": row.get("Message ID", ""),
            "conversation_id": row.get("Conversation ID", ""),
            "received": row.get("Received", ""),
            "sender_email": row.get("Sender Email", ""),
            "subject": row.get("Subject", ""),
            "category": row.get("Category", ""),
            "confidence": row.get("Confidence", ""),
            "run": run_meta or {},
        }
    save_json_dict(path, log)
    return added



def format_date(value: dt.datetime) -> str:
    return value.astimezone().strftime("%Y-%m-%d")


def guess_company(sender_name: str, sender_email: str, subject: str) -> str:
    domain = sender_domain(sender_email)
    if domain:
        base = domain.split(".")[0]
        bad = {"mail", "email", "noreply", "no-reply", "notifications", "jobs", "careers", "talent"}
        if base not in bad and len(base) > 2:
            return base.replace("-", " ").title()
    if sender_name:
        cleaned = re.sub(
            r"\b(no.?reply|notification|jobs|careers|talent acquisition|recruiting)\b",
            "",
            sender_name,
            flags=re.I,
        ).strip(" -|:")
        if cleaned:
            return cleaned[:60]
    match = re.search(r"(?:at|with)\s+([A-Z][A-Za-z0-9& .'-]{2,50})", subject or "")
    return match.group(1).strip() if match else ""


def make_followup(record: EmailRecord, classification_hint: str, company: str, config: dict) -> Tuple[str, str]:
    your_name = config["user"].get("your_name", "Your Name")
    signature = config["user"].get("signature") or f"Best,\n{your_name}"
    company_label = company or "there"
    clean_subject = record.subject or "my application"
    if len(clean_subject) > 90:
        clean_subject = clean_subject[:87] + "..."
    subject = f"Following up: {clean_subject}"
    if classification_hint == "Application Receipt":
        body = (
            f"Hi {company_label},\n\n"
            "I hope you're doing well. I wanted to follow up on my application and see if there are any updates on next steps. "
            "I'm still very interested and would appreciate any information when you have a chance.\n\n"
            f"{signature}"
        )
    elif classification_hint == "Recruiter / Important":
        body = (
            f"Hi {company_label},\n\n"
            "I hope you're doing well. I wanted to follow up on this and check whether there are any updates or next steps from my side.\n\n"
            f"{signature}"
        )
    else:
        body = f"Hi {company_label},\n\nI hope you're doing well. I wanted to follow up and check in on this.\n\n{signature}"
    return subject, body


def classify_email_rules(record: EmailRecord, config: dict) -> Classification:
    """Local classifier.

    Priority is intentionally conservative:
      1. Rejections override everything else.
      2. True action requests only when action phrasing is specific.
      3. Receipts become Applications or Follow Up Queue after the wait period.
      4. Job-board spam only after human/job signals are considered.

    This prevents normal receipt emails that mention vague "next steps" or "review"
    from being pushed into Needs Action.
    """
    c = config["classification"]
    text = f"{record.sender_name} {record.sender_email} {record.subject} {record.preview} {record.body_text[:1800]}"
    text_n = norm(text)
    domain = sender_domain(record.sender_email)
    received_dt = safe_date(record.received)
    today = dt.datetime.now(dt.timezone.utc)
    age_days = max(0, (today - received_dt).days)
    company = guess_company(record.sender_name, record.sender_email, record.subject)

    for ignored in c.get("ignore_domains", []):
        if domain.endswith(ignored.lower()):
            return Classification(
                category="Ignore / Non-Job",
                confidence=0.95,
                reason=f"Sender domain matches ignored domain: {ignored}",
                suggested_action="Ignore unless you personally need it.",
                follow_up_date="",
                approved_to_send="NO",
                follow_up_subject="",
                follow_up_body="",
                target_folder_key="",
                move_eligible=False,
                company_guess=company,
            )

    has_rejection, rejection_hits = contains_any(text_n, c["rejection_phrases"])
    has_action, action_hits = contains_any(text_n, c["important_action_phrases"])
    has_receipt, receipt_hits = contains_any(text_n, c["application_receipt_phrases"])
    has_job, job_hits = contains_any(text_n, c["job_relevant_phrases"])
    has_spam_phrase, spam_hits = contains_any(text_n, c["spam_sender_phrases"])

    is_board_domain = any(domain.endswith(d.lower()) for d in c["job_board_domains"])
    is_trusted_human = any(domain.endswith(d.lower()) for d in c.get("trusted_human_domains", []))

    # Rejections should win over generic "application" or "next steps" wording.
    if has_rejection and (has_job or is_board_domain or "application" in text_n or "candidate" in text_n):
        confidence = 0.80 + min(0.15, 0.03 * len(rejection_hits))
        return Classification(
            category="Rejection",
            confidence=min(confidence, 0.98),
            reason="Rejection phrase found: " + ", ".join(rejection_hits[:3]),
            suggested_action="Move to Rejections. No follow-up needed.",
            follow_up_date="",
            approved_to_send="NO",
            follow_up_subject="",
            follow_up_body="",
            target_folder_key="rejections",
            move_eligible=True,
            company_guess=company,
        )

    # Bare "assessment" / "next steps" should not count. Keep this check specific.
    specific_action_patterns = [
        r"\b(select|choose|pick|book|schedule)\s+(a\s+)?(time|slot|interview|call|meeting)\b",
        r"\b(reply|respond)\s+(with|to)\s+(your\s+)?availability\b",
        r"\b(are|would)\s+you\s+available\b",
        r"\b(phone|video)\s+(screen|interview|call)\b",
        r"\b(calendly|calendar invite)\b",
        r"\b(action required|please complete|required to complete)\b",
        r"\b(complete|take|submit)\s+(this\s+|the\s+)?(assessment|test|task|form)\b",
        r"\bwe(\'|’)d like to speak\b",
        r"\bwe would like to speak\b",
        r"\binterview invitation\b",
        r"\binvite you to interview\b",
        r"\boffer\b",
    ]
    has_specific_action = bool(has_action and any(re.search(p, text_n, flags=re.I) for p in specific_action_patterns))

    if has_specific_action:
        wait_days = int(config["follow_up"].get("recruiter_wait_days", 3))
        due = today if age_days >= wait_days else received_dt + dt.timedelta(days=wait_days)
        subject, body = make_followup(record, "Recruiter / Important", company, config)
        confidence = 0.88 + min(0.08, 0.02 * len(action_hits))
        return Classification(
            category="Needs Action",
            confidence=min(confidence, 0.98),
            reason="Specific action/interview phrase found: " + ", ".join(action_hits[:3]),
            suggested_action="Review now. This may require scheduling, replying, or completing an assessment.",
            follow_up_date=format_date(due),
            approved_to_send="NO",
            follow_up_subject=subject,
            follow_up_body=body,
            target_folder_key="important",
            move_eligible=False,
            company_guess=company,
        )

    # Receipts should be classified before job-board spam if they are from ATS systems.
    if has_receipt:
        wait_days = int(config["follow_up"].get("application_receipt_wait_days", 7))
        due = received_dt + dt.timedelta(days=wait_days)
        subject, body = make_followup(record, "Application Receipt", company, config)
        category = "Follow Up Queue" if due <= today else "Application Receipt"
        return Classification(
            category=category,
            confidence=0.82,
            reason="Application receipt phrase found: " + ", ".join(receipt_hits[:3]),
            suggested_action=(
                "Follow up now; the waiting period has passed."
                if due <= today
                else f"Follow up after {format_date(due)} if there is no response."
            ),
            follow_up_date=format_date(due),
            approved_to_send="NO",
            follow_up_subject=subject,
            follow_up_body=body,
            target_folder_key="follow_up",
            move_eligible=False,
            company_guess=company,
        )

    if is_board_domain and has_spam_phrase and not is_trusted_human:
        confidence = 0.82 + min(0.12, 0.03 * len(spam_hits))
        return Classification(
            category="Job Board Spam",
            confidence=min(confidence, 0.98),
            reason=f"Job-board domain + alert phrasing: {domain}; " + ", ".join(spam_hits[:3]),
            suggested_action="Move to Job Board Spam. These are usually automated listing blasts.",
            follow_up_date="",
            approved_to_send="NO",
            follow_up_subject="",
            follow_up_body="",
            target_folder_key="job_board_spam",
            move_eligible=True,
            company_guess=company,
        )

    if has_job and not (is_board_domain and has_spam_phrase):
        wait_days = int(config["follow_up"].get("default_wait_days", 5))
        due = received_dt + dt.timedelta(days=wait_days)
        subject, body = make_followup(record, "Recruiter / Important", company, config)
        category = "Important Job Email" if age_days < wait_days else "Follow Up Queue"
        return Classification(
            category=category,
            confidence=0.64,
            reason="General job-search wording found: " + ", ".join(job_hits[:3]),
            suggested_action="Review. This may be relevant to your job search.",
            follow_up_date=format_date(due),
            approved_to_send="NO",
            follow_up_subject=subject,
            follow_up_body=body,
            target_folder_key="follow_up" if category == "Follow Up Queue" else "important",
            move_eligible=False,
            company_guess=company,
        )

    if is_board_domain:
        return Classification(
            category="Job Board Spam",
            confidence=0.68,
            reason=f"Sender is a known job-board domain: {domain}",
            suggested_action="Likely automated. Review once before moving if confidence feels low.",
            follow_up_date="",
            approved_to_send="NO",
            follow_up_subject="",
            follow_up_body="",
            target_folder_key="job_board_spam",
            move_eligible=True,
            company_guess=company,
        )

    return Classification(
        category="Ignore / Non-Job",
        confidence=0.50,
        reason="No strong job-search signal found.",
        suggested_action="No action.",
        follow_up_date="",
        approved_to_send="NO",
        follow_up_subject="",
        follow_up_body="",
        target_folder_key="",
        move_eligible=False,
        company_guess=company,
    )


def extract_openrouter_text(data: dict) -> str:
    choices = data.get("choices") or []
    if not choices:
        raise ValueError("OpenRouter returned no choices")
    message = choices[0].get("message") or {}
    content = message.get("content", "")
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict) and item.get("type") == "text":
                parts.append(str(item.get("text", "")))
        return "\n".join(parts)
    return str(content)


def parse_json_object(text: str) -> dict:
    text = (text or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.S)
        if not match:
            raise
        return json.loads(match.group(0))


def classification_from_ai_payload(payload: dict, fallback: Classification, model: str, cached: bool = False) -> Classification:
    category = str(payload.get("category") or fallback.category).strip()
    if category not in ALLOWED_CATEGORIES:
        category = fallback.category
    try:
        confidence = float(payload.get("confidence", fallback.confidence))
    except (TypeError, ValueError):
        confidence = fallback.confidence
    confidence = max(0.0, min(1.0, confidence))
    target_folder_key = str(payload.get("target_folder_key") or fallback.target_folder_key or "").strip()
    if target_folder_key not in {"rejections", "job_board_spam", "follow_up", "important", ""}:
        target_folder_key = fallback.target_folder_key
    move_eligible = bool(payload.get("move_eligible", fallback.move_eligible))
    if category not in {"Rejection", "Job Board Spam"}:
        move_eligible = False
    if category == "Rejection":
        target_folder_key = "rejections"
    elif category == "Job Board Spam":
        target_folder_key = "job_board_spam"
    elif category in {"Follow Up Queue", "Application Receipt"}:
        target_folder_key = "follow_up"
    elif category in {"Needs Action", "Important Job Email"}:
        target_folder_key = "important"
    else:
        target_folder_key = ""
    approved = str(payload.get("approved_to_send") or "NO").upper().strip()
    if approved != "YES":
        approved = "NO"
    # Safety gate: the AI is never allowed to pre-approve sending.
    approved = "NO"
    return Classification(
        category=category,
        confidence=confidence,
        reason=str(payload.get("reason") or fallback.reason)[:600],
        suggested_action=str(payload.get("suggested_action") or fallback.suggested_action)[:600],
        follow_up_date=str(payload.get("follow_up_date") or fallback.follow_up_date or "")[:20],
        approved_to_send=approved,
        follow_up_subject=str(payload.get("follow_up_subject") or fallback.follow_up_subject or "")[:240],
        follow_up_body=str(payload.get("follow_up_body") or fallback.follow_up_body or "")[:4000],
        target_folder_key=target_folder_key,
        move_eligible=move_eligible,
        company_guess=str(payload.get("company_guess") or fallback.company_guess or "")[:120],
        ai_used=True,
        ai_model=f"{model}{' cached' if cached else ''}",
    )


def get_sender(message: dict) -> Tuple[str, str]:
    sender = message.get("from") or message.get("sender") or {}
    addr = sender.get("emailAddress") or {}
    return addr.get("name", ""), addr.get("address", "")


def fetch_recent_messages(
    client: GraphClient,
    source: str,
    days_back: int,
    max_messages: int,
    page_size: int,
    include_body_text: bool,
) -> List[EmailRecord]:
    start = (dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=days_back)).replace(microsecond=0)
    filter_value = start.isoformat().replace("+00:00", "Z")
    fields = [
        "id",
        "conversationId",
        "internetMessageId",
        "subject",
        "receivedDateTime",
        "from",
        "sender",
        "bodyPreview",
        "webLink",
        "importance",
        "hasAttachments",
    ]
    if include_body_text:
        fields.append("body")
    select = ",".join(fields)
    params = urlencode(
        {
            "$top": min(page_size, 100),
            "$select": select,
            "$orderby": "receivedDateTime desc",
            "$filter": f"receivedDateTime ge {filter_value}",
        }
    )
    if source == "all":
        endpoint = f"/me/messages?{params}"
    else:
        endpoint = f"/me/mailFolders/inbox/messages?{params}"
    headers = {"Prefer": 'outlook.body-content-type="text"'} if include_body_text else {}
    raw_messages = client.paged_get(endpoint, limit=max_messages, headers=headers)
    records: List[EmailRecord] = []
    for msg in raw_messages:
        name, email_addr = get_sender(msg)
        body = msg.get("body") or {}
        body_text = body.get("content", "") if isinstance(body, dict) else ""
        records.append(
            EmailRecord(
                id=msg.get("id", ""),
                conversation_id=msg.get("conversationId", ""),
                subject=msg.get("subject", ""),
                received=msg.get("receivedDateTime", ""),
                sender_name=name,
                sender_email=email_addr,
                preview=msg.get("bodyPreview", ""),
                body_text=strip_html(body_text),
                web_link=msg.get("webLink", ""),
                internet_message_id=msg.get("internetMessageId", ""),
                importance=msg.get("importance", "normal"),
                has_attachments=bool(msg.get("hasAttachments", False)),
            )
        )
    return records


def ensure_folder_path(client: GraphClient, root_name: str, child_name: str) -> str:
    def list_children(parent_id: Optional[str]) -> List[dict]:
        if parent_id:
            return client.paged_get(f"/me/mailFolders/{parent_id}/childFolders?$top=100")
        return client.paged_get("/me/mailFolders?$top=100")

    def find_or_create(parent_id: Optional[str], name: str) -> str:
        children = list_children(parent_id)
        for folder in children:
            if folder.get("displayName", "").lower() == name.lower():
                return folder["id"]
        body = {"displayName": name}
        if parent_id:
            created = client.request("POST", f"/me/mailFolders/{parent_id}/childFolders", json=body)
        else:
            created = client.request("POST", "/me/mailFolders", json=body)
        return created["id"]

    root_id = find_or_create(None, root_name)
    return find_or_create(root_id, child_name)


def move_message(client: GraphClient, message_id: str, destination_folder_id: str) -> dict:
    return client.request("POST", f"/me/messages/{message_id}/move", json={"destinationId": destination_folder_id})


def create_draft(client: GraphClient, to_email: str, subject: str, body: str) -> dict:
    payload = {
        "subject": subject,
        "importance": "normal",
        "body": {"contentType": "Text", "content": body},
        "toRecipients": [{"emailAddress": {"address": to_email}}],
    }
    return client.request("POST", "/me/messages", json=payload)


def send_mail(client: GraphClient, to_email: str, subject: str, body: str) -> None:
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "Text", "content": body},
            "toRecipients": [{"emailAddress": {"address": to_email}}],
        },
        "saveToSentItems": True,
    }
    client.request("POST", "/me/sendMail", json=payload)


def row_from(record: EmailRecord, cls: Classification) -> Dict[str, Any]:
    received_dt = safe_date(record.received).astimezone()
    return {
        "Message ID": record.id,
        "Processed Key": message_fingerprint(record),
        "Conversation ID": record.conversation_id,
        "Received": received_dt.strftime("%Y-%m-%d %H:%M"),
        "Sender Name": record.sender_name,
        "Sender Email": record.sender_email,
        "Company Guess": cls.company_guess,
        "Subject": record.subject,
        "Category": cls.category,
        "Confidence": round(cls.confidence, 2),
        "AI Used": "YES" if cls.ai_used else "NO",
        "AI Model": cls.ai_model,
        "AI Error": cls.ai_error,
        "Reason": cls.reason,
        "Suggested Action": cls.suggested_action,
        "Follow Up Date": cls.follow_up_date,
        "Approved To Send": cls.approved_to_send,
        "Follow Up Subject": cls.follow_up_subject,
        "Follow Up Body": cls.follow_up_body,
        "Moved?": "NO",
        "Send Status": "",
        "Outlook Link": record.web_link,
        "Preview": record.preview,
    }


def setup_sheet(ws, title_fill="1F4E78") -> None:
    header_fill = PatternFill("solid", fgColor=title_fill)
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def append_table(ws, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def add_approved_dropdown(ws, headers: List[str]) -> None:
    if "Approved To Send" not in headers:
        return
    col = headers.index("Approved To Send") + 1
    letter = get_column_letter(col)
    dv = DataValidation(type="list", formula1='"NO,YES"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{max(ws.max_row, 2)}")


def style_workbook(wb: Workbook) -> None:
    widths = {
        "A": 24,
        "B": 17,
        "C": 22,
        "D": 28,
        "E": 18,
        "F": 45,
        "G": 18,
        "H": 11,
        "I": 10,
        "J": 22,
        "K": 26,
        "L": 25,
        "M": 36,
        "N": 40,
        "O": 14,
        "P": 16,
        "Q": 38,
        "R": 55,
        "S": 12,
        "T": 18,
        "U": 18,
        "V": 55,
    }
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            setup_sheet(ws)
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].height = 45
        headers = [cell.value for cell in ws[1]] if ws.max_row else []
        if "Approved To Send" in headers:
            col = headers.index("Approved To Send") + 1
            letter = get_column_letter(col)
            ws.conditional_formatting.add(
                f"{letter}2:{letter}{max(ws.max_row, 2)}",
                FormulaRule(formula=[f'${letter}2="YES"'], fill=PatternFill("solid", fgColor="FFF2CC")),
            )
            add_approved_dropdown(ws, headers)
        if "Outlook Link" in headers:
            link_col = headers.index("Outlook Link") + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row_idx, link_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.value = "Open in Outlook"
                    cell.style = "Hyperlink"


def write_report(rows: List[Dict[str, Any]], output_path: Path) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    headers = [
        "Message ID",
        "Processed Key",
        "Received",
        "Sender Name",
        "Sender Email",
        "Company Guess",
        "Subject",
        "Category",
        "Confidence",
        "AI Used",
        "AI Model",
        "AI Error",
        "Reason",
        "Suggested Action",
        "Follow Up Date",
        "Approved To Send",
        "Follow Up Subject",
        "Follow Up Body",
        "Moved?",
        "Send Status",
        "Outlook Link",
        "Preview",
        "Conversation ID",
    ]
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    counts = Counter(row["Category"] for row in rows)
    summary.append(["Total scanned", len(rows)])
    summary.append(["AI-classified rows", sum(1 for row in rows if row.get("AI Used") == "YES")])
    for category, count in counts.most_common():
        summary.append([category, count])
    summary.append(["", ""])
    summary.append(["Safety", "Sending is never automatic from review/apply. Change Approved To Send to YES, then run draft-approved or send-approved."])
    summary.append(["Safety", "apply mode moves high-confidence Rejections and Job Board Spam only; it does not permanently delete messages."])
    sheet_rules = [
        ("Action Needed", lambda r: r["Category"] == "Needs Action"),
        ("Follow Up Queue", lambda r: r["Category"] == "Follow Up Queue"),
        ("Applications", lambda r: r["Category"] == "Application Receipt"),
        ("Rejections", lambda r: r["Category"] == "Rejection"),
        ("Job Board Spam", lambda r: r["Category"] == "Job Board Spam"),
        ("Important", lambda r: r["Category"] == "Important Job Email"),
        ("All Scanned", lambda r: True),
    ]
    for sheet_name, predicate in sheet_rules:
        ws = wb.create_sheet(sheet_name)
        selected = [row for row in rows if predicate(row)]
        append_table(ws, selected, headers)
    instructions = wb.create_sheet("Instructions")
    instructions.append(["Step", "What to do"])
    instructions.append(["1", "Open Action Needed and Follow Up Queue first."])
    instructions.append(["2", "Read the AI-written Follow Up Body and edit it if needed."])
    instructions.append(["3", "Only change Approved To Send to YES after you personally review the body."])
    instructions.append(["4", "Run draft-approved to create Outlook drafts, or send-approved --yes-really to send."])
    instructions.append(["5", "Run apply when you want high-confidence Rejections and Job Board Spam moved to Job Hunt folders."])
    instructions.append(["Privacy", "When OpenRouter is enabled, email subject/sender/body text is sent to the configured model provider through OpenRouter."])
    style_workbook(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def apply_moves(client: GraphClient, rows: List[Dict[str, Any]], config: dict) -> None:
    folder_cfg = config["folders"]
    auto = config["automation"]
    root = folder_cfg["root"]
    min_conf = float(auto.get("min_confidence_to_move", 0.72))
    folder_cache: Dict[str, str] = {}
    key_to_child = {
        "rejections": folder_cfg["rejections"],
        "job_board_spam": folder_cfg["job_board_spam"],
        "follow_up": folder_cfg["follow_up"],
        "important": folder_cfg["important"],
    }

    def folder_for_key(key: str) -> str:
        if key not in folder_cache:
            folder_cache[key] = ensure_folder_path(client, root, key_to_child[key])
        return folder_cache[key]

    for row in rows:
        category = row["Category"]
        confidence = float(row.get("Confidence") or 0)
        folder_key = None
        if category == "Rejection" and auto.get("auto_move_rejections", True):
            folder_key = "rejections"
        elif category == "Job Board Spam" and auto.get("auto_move_job_board_spam", True):
            folder_key = "job_board_spam"
        elif category in ("Follow Up Queue", "Application Receipt") and auto.get("auto_move_follow_up", False):
            folder_key = "follow_up"
        if not folder_key or confidence < min_conf:
            continue
        try:
            move_message(client, row["Message ID"], folder_for_key(folder_key))
            row["Moved?"] = f"YES → {root}/{key_to_child[folder_key]}"
            print(f"Moved: {category} | {row['Subject'][:70]}")
        except Exception as exc:
            row["Moved?"] = f"ERROR: {exc}"
            print(f"Move failed for {row['Subject'][:70]}: {exc}")


def load_sent_log(path: Path) -> dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def save_sent_log(path: Path, data: dict) -> None:
    path.write_text(json.dumps(data, indent=2, sort_keys=True), encoding="utf-8")


def parse_report_rows(report_path: Path, sheet_names: List[str]) -> Tuple[Workbook, List[Tuple[str, int, Dict[str, Any]]]]:
    wb = load_workbook(report_path)
    rows: List[Tuple[str, int, Dict[str, Any]]] = []
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        for row_idx in range(2, ws.max_row + 1):
            data = {headers[col_idx - 1]: ws.cell(row_idx, col_idx).value for col_idx in range(1, len(headers) + 1)}
            if any(data.values()):
                rows.append((sheet_name, row_idx, data))
    return wb, rows


def is_due(date_text: Any) -> bool:
    if not date_text:
        return False
    if isinstance(date_text, dt.datetime):
        d = date_text.date()
    elif isinstance(date_text, dt.date):
        d = date_text
    else:
        try:
            d = dt.date.fromisoformat(str(date_text)[:10])
        except ValueError:
            return False
    return d <= dt.datetime.now().date()


def process_approved_messages(client: GraphClient, report_path: Path, mode: str, config: dict, yes_really: bool) -> None:
    if mode == "send" and not yes_really:
        raise SystemExit("Refusing to send email without --yes-really. Use draft-approved for a safer first pass.")
    wb, rows = parse_report_rows(report_path, ["Action Needed", "Follow Up Queue", "Important", "Applications"])
    log_path = report_path.parent / SENT_LOG_FILE
    sent_log = load_sent_log(log_path)
    send_only_if_due = bool(config["follow_up"].get("send_only_if_due", True))
    total = 0
    for sheet_name, row_idx, row in rows:
        approved = str(row.get("Approved To Send") or "").strip().upper()
        status = str(row.get("Send Status") or "").strip()
        if approved != "YES" or status:
            continue
        if send_only_if_due and not is_due(row.get("Follow Up Date")):
            continue
        to_email = row.get("Sender Email")
        subject = row.get("Follow Up Subject") or f"Following up: {row.get('Subject') or ''}"
        body = row.get("Follow Up Body") or ""
        if not to_email or not body:
            continue
        if is_no_reply_address(str(to_email)):
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if "Send Status" in headers:
                ws.cell(row_idx, headers.index("Send Status") + 1).value = "Skipped: sender is no-reply/ATS; find a real recruiter/contact first"
            print(f"Skipped no-reply/ATS sender: {to_email} | {subject[:70]}")
            continue
        unique_key = f"{row.get('Message ID')}|{to_email}|{subject}"
        if unique_key in sent_log:
            continue
        if mode == "draft":
            result = create_draft(client, to_email, subject, body)
            sent_log[unique_key] = {"mode": "draft", "when": dt.datetime.now().isoformat(), "draft_id": result.get("id")}
            new_status = f"Draft created {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            print(f"Drafted: {to_email} | {subject[:70]}")
        else:
            send_mail(client, to_email, subject, body)
            sent_log[unique_key] = {"mode": "sent", "when": dt.datetime.now().isoformat()}
            new_status = f"Sent {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
            print(f"Sent: {to_email} | {subject[:70]}")
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        if "Send Status" in headers:
            ws.cell(row_idx, headers.index("Send Status") + 1).value = new_status
        total += 1
    wb.save(report_path)
    save_sent_log(log_path, sent_log)
    print(f"Done. Processed {total} approved follow-up(s).")


def maybe_get_ai_agent(config: dict, project_dir: Path, force_no_ai: bool) -> Optional[OpenRouterAgent]:
    if force_no_ai:
        return None
    ai_cfg = config.get("ai", {})
    if not ai_cfg.get("enabled", True):
        return None
    if not ai_cfg.get("send_email_content_to_openrouter", True):
        print("AI is enabled but send_email_content_to_openrouter is false, so using local rules only.")
        return None
    api_key = os.getenv(ai_cfg.get("api_key_env", "OPENROUTER_API_KEY"), "").strip()
    if not api_key:
        print("No OPENROUTER_API_KEY found. Using local rule-based sorting only.")
        return None
    model_env = ai_cfg.get("model_env", "OPENROUTER_MODEL")
    model = os.getenv(model_env, ai_cfg.get("default_model", "~openai/gpt-latest")).strip()
    print(f"OpenRouter AI enabled with model: {model}")
    return OpenRouterAgent(api_key=api_key, model=model, config=config, cache_path=project_dir / AI_CACHE_FILE)


def should_call_ai(record: EmailRecord, rule_cls: Classification, config: dict) -> bool:
    if not config.get("ai", {}).get("only_ai_for_job_likely_messages", False):
        return True
    if rule_cls.category != "Ignore / Non-Job":
        return True
    text = norm(f"{record.sender_name} {record.sender_email} {record.subject} {record.preview}")
    has_job, _ = contains_any(text, config["classification"].get("job_relevant_phrases", []))
    return has_job


def reconcile_classification(record: EmailRecord, rule_cls: Classification, cls: Classification) -> Classification:
    """Apply local safety priority after optional AI classification.

    The AI can add nuance, but it should not turn obvious rejections into follow-ups
    or ordinary receipts into Needs Action unless the local rules found a specific action.
    """
    if rule_cls.category == "Rejection" and rule_cls.confidence >= 0.75:
        return dataclasses.replace(
            rule_cls,
            ai_used=cls.ai_used,
            ai_model=cls.ai_model,
            ai_error=cls.ai_error or ("AI result overridden by local rejection safety rule." if cls.category != "Rejection" else ""),
        )

    if rule_cls.category in {"Application Receipt", "Follow Up Queue"} and cls.category == "Needs Action":
        # Keep receipt/follow-up classification unless the local rules independently found a specific action.
        return dataclasses.replace(
            rule_cls,
            ai_used=cls.ai_used,
            ai_model=cls.ai_model,
            ai_error=cls.ai_error or "AI Needs Action overridden because local rules found an application receipt, not a specific action request.",
        )

    if cls.category in {"Follow Up Queue", "Application Receipt", "Needs Action", "Important Job Email"} and is_no_reply_address(record.sender_email):
        return dataclasses.replace(
            cls,
            suggested_action=(cls.suggested_action + " Sender appears to be a no-reply/ATS address; find a real recruiter/contact before following up.").strip(),
            follow_up_body="",
            follow_up_subject="",
            ai_error=(cls.ai_error + " " if cls.ai_error else "") + "No-reply sender safety: follow-up draft suppressed.",
        )

    return cls


def build_rows(records: List[EmailRecord], config: dict, agent: Optional[OpenRouterAgent]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    include_non_job = bool(config["scan"].get("include_non_job_related_in_all_sheet", True))
    min_override = float(config.get("ai", {}).get("min_ai_confidence_to_override_rules", 0.55))
    ai_attempted = 0
    ai_failed = 0
    ai_overridden_low_conf = 0
    ai_safety_overridden = 0

    for idx, record in enumerate(records, start=1):
        rule_cls = classify_email_rules(record, config)
        cls = rule_cls
        if agent and should_call_ai(record, rule_cls, config):
            ai_attempted += 1
            try:
                ai_cls = agent.classify(record, rule_cls)
                if ai_cls.confidence >= min_override:
                    cls = ai_cls
                else:
                    ai_overridden_low_conf += 1
                    cls = dataclasses.replace(
                        rule_cls,
                        ai_used=True,
                        ai_model=agent.model,
                        ai_error=f"AI confidence {ai_cls.confidence:.2f} below override threshold; used local rule result.",
                    )
            except Exception as exc:
                ai_failed += 1
                if config.get("ai", {}).get("fallback_to_rules_if_ai_fails", True):
                    cls = dataclasses.replace(rule_cls, ai_error=str(exc)[:500])
                else:
                    raise

        before_reconcile = cls.category
        cls = reconcile_classification(record, rule_cls, cls)
        if before_reconcile != cls.category or ("overridden" in (cls.ai_error or "").lower()):
            ai_safety_overridden += 1

        if cls.category == "Ignore / Non-Job" and not include_non_job:
            continue
        rows.append(row_from(record, cls))
        if idx % 10 == 0:
            print(f"Classified {idx}/{len(records)} messages...")

    print(
        "Classification run stats: "
        f"records={len(records)}, report_rows={len(rows)}, "
        f"ai_attempted={ai_attempted}, ai_failed={ai_failed}, "
        f"ai_low_confidence_overrides={ai_overridden_low_conf}, "
        f"safety_overrides={ai_safety_overridden}"
    )
    return rows


def get_client(config: dict, project_dir: Path) -> GraphClient:
    load_dotenv(project_dir / ".env")
    client_id = os.getenv("MS_CLIENT_ID", "").strip()
    if not client_id:
        raise SystemExit("Missing MS_CLIENT_ID. Put it in .env or set it as an environment variable.")
    tenant = os.getenv("MS_TENANT", config["graph"].get("tenant", "consumers")).strip() or "consumers"
    scopes = os.getenv("MS_SCOPES")
    raw_scope_list = [s.strip() for s in scopes.split(",")] if scopes else list(config["graph"].get("scopes", DEFAULT_SCOPES))

    scope_list: List[str] = []
    removed_reserved: List[str] = []
    for raw in raw_scope_list:
        scope = str(raw or "").strip()
        if not scope:
            continue
        if scope in RESERVED_MSAL_SCOPES:
            removed_reserved.append(scope)
            continue
        scope_list.append(scope)

    if removed_reserved:
        print(f"Ignoring MSAL-reserved scope(s): {', '.join(sorted(set(removed_reserved)))}")
    if not scope_list:
        raise SystemExit("No valid Microsoft Graph scopes configured. Use User.Read,Mail.ReadWrite,Mail.Send.")

    return GraphClient(client_id=client_id, tenant=tenant, scopes=scope_list, cache_path=project_dir / TOKEN_CACHE_FILE)

def run_review_or_apply(args: argparse.Namespace, config: dict, project_dir: Path) -> None:
    load_dotenv(project_dir / ".env")
    client = get_client(config, project_dir)
    days_back = args.days or int(config["scan"].get("days_back", 60))
    max_messages = args.max or int(config["scan"].get("max_messages", 250))
    page_size = int(config["graph"].get("page_size", 50))
    source = args.source or config["scan"].get("source", "inbox")
    include_body_text = bool(config["scan"].get("include_body_text", True))
    print(f"Scanning {source}: last {days_back} days, up to {max_messages} messages...")
    records = fetch_recent_messages(
        client,
        source=source,
        days_back=days_back,
        max_messages=max_messages,
        page_size=page_size,
        include_body_text=include_body_text,
    )

    processed_cfg = config.get("processed", {})
    processed_enabled = bool(processed_cfg.get("enabled", True))
    skip_processed = bool(getattr(args, "skip_processed", False) or processed_cfg.get("skip_previously_processed", False))
    mark_processed = bool(getattr(args, "mark_processed", False) or processed_cfg.get("mark_as_processed_after_report", False))
    processed_path = processed_log_path(project_dir, config)

    fetched_count = len(records)
    skipped = 0
    if processed_enabled and skip_processed:
        processed_log = load_json_dict(processed_path)
        before = len(records)
        records = [record for record in records if message_fingerprint(record) not in processed_log]
        skipped = before - len(records)
        print(f"Skipped {skipped} previously processed message(s) using {processed_path.name}.")
    elif processed_enabled:
        print("Processed-message skipping is OFF for this run. Use --skip-processed or set it in config.yaml to avoid rescanning.")

    print(f"Fetched messages: {fetched_count}")
    print(f"New/unprocessed messages to classify: {len(records)}")

    agent = maybe_get_ai_agent(config, project_dir, force_no_ai=args.no_ai)
    rows = build_rows(records, config, agent)
    if args.command == "apply":
        print("Applying safe moves for high-confidence Rejections and Job Board Spam...")
        apply_moves(client, rows, config)
    out_path = Path(args.output) if args.output else project_dir / "output" / "job_hunt_email_report.xlsx"
    write_report(rows, out_path)
    if processed_enabled and mark_processed:
        added = mark_rows_processed(
            processed_path,
            rows,
            run_meta={
                "command": args.command,
                "days": days_back,
                "max": max_messages,
                "source": source,
                "report": str(out_path),
            },
        )
        print(f"Marked {added} new message(s) as processed in: {processed_path}")
    counts = Counter(row["Category"] for row in rows)
    print(f"\nReport written to: {out_path}")
    print("Summary:")
    print(f"  Fetched: {fetched_count}")
    print(f"  Skipped previously processed: {skipped}")
    print(f"  Classified this run: {len(records)}")
    print(f"  Rows written to report: {len(rows)}")
    for category, count in counts.most_common():
        print(f"  {category}: {count}")


def write_env_template(path: Path) -> None:
    if path.exists():
        return
    path.write_text(
        "# Copy this file to .env and fill in the values.\n"
        "\n"
        "# Microsoft Graph / Azure App Registration\n"
        "MS_CLIENT_ID=PASTE-YOUR-APPLICATION-CLIENT-ID-HERE\n"
        "# For personal Outlook/Hotmail, use consumers. For mixed work/personal, use common.\n"
        "MS_TENANT=consumers\n"
        "# Optional override; comma-separated. Usually leave alone.\n"
        "# MS_SCOPES=User.Read,Mail.ReadWrite,Mail.Send\n"
        "\n"
        "# OpenRouter AI language/classification agent\n"
        "OPENROUTER_API_KEY=PASTE-YOUR-OPENROUTER-KEY-HERE\n"
        "# Pick any model slug from OpenRouter. The docs also support latest aliases like ~openai/gpt-latest.\n"
        "OPENROUTER_MODEL=openrouter/auto\n"
        "# Optional attribution headers for OpenRouter dashboards/leaderboards\n"
        "OPENROUTER_APP_NAME=Job Hunt Email Agent\n"
        "# OPENROUTER_SITE_URL=https://example.com\n",
        encoding="utf-8",
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Job Hunt Outlook Email Agent + OpenRouter AI")
    parser.add_argument("--config", default="config.yaml", help="Path to config.yaml")
    sub = parser.add_subparsers(dest="command", required=True)

    p_review = sub.add_parser("review", help="Create Excel report only. No mailbox changes.")
    p_review.add_argument("--days", type=int, help="Days back to scan")
    p_review.add_argument("--max", type=int, help="Maximum messages to scan")
    p_review.add_argument("--source", choices=["inbox", "all"], help="Scan inbox only or all messages")
    p_review.add_argument("--output", help="Output .xlsx path")
    p_review.add_argument("--no-ai", action="store_true", help="Disable OpenRouter and use local rules only")
    p_review.add_argument("--skip-processed", action="store_true", help="Skip emails already listed in processed_messages.json")
    p_review.add_argument("--mark-processed", action="store_true", help="After writing the report, add those rows to processed_messages.json")

    p_apply = sub.add_parser("apply", help="Create report and move high-confidence low-risk categories.")
    p_apply.add_argument("--days", type=int, help="Days back to scan")
    p_apply.add_argument("--max", type=int, help="Maximum messages to scan")
    p_apply.add_argument("--source", choices=["inbox", "all"], help="Scan inbox only or all messages")
    p_apply.add_argument("--output", help="Output .xlsx path")
    p_apply.add_argument("--no-ai", action="store_true", help="Disable OpenRouter and use local rules only")
    p_apply.add_argument("--skip-processed", action="store_true", help="Skip emails already listed in processed_messages.json")
    p_apply.add_argument("--mark-processed", action="store_true", help="After writing the report, add those rows to processed_messages.json")

    p_mark = sub.add_parser("mark-report-processed", help="Mark all rows in an existing Excel report as processed.")
    p_mark.add_argument("--report", required=True, help="Path to job_hunt_email_report.xlsx")

    sub.add_parser("processed-status", help="Show how many emails are in processed_messages.json.")
    sub.add_parser("reset-processed", help="Delete processed_messages.json so old emails can be scanned again.")

    p_send = sub.add_parser("send-approved", help="Send only approved follow-ups from the Excel report.")
    p_send.add_argument("--report", required=True, help="Path to job_hunt_email_report.xlsx")
    p_send.add_argument("--yes-really", action="store_true", help="Required to actually send emails")

    p_draft = sub.add_parser("draft-approved", help="Create drafts for approved follow-ups from the Excel report.")
    p_draft.add_argument("--report", required=True, help="Path to job_hunt_email_report.xlsx")

    sub.add_parser("init", help="Create config.yaml and .env.template")

    args = parser.parse_args()
    project_dir = Path.cwd()
    config_path = Path(args.config)

    if args.command == "init":
        save_default_config(config_path)
        write_env_template(project_dir / ".env.template")
        print(f"Created/checked {config_path} and {project_dir / '.env.template'}")
        return

    config = load_config(config_path)
    if args.command in {"review", "apply"}:
        run_review_or_apply(args, config, project_dir)

    elif args.command == "mark-report-processed":
        _, parsed = parse_report_rows(
            Path(args.report),
            ["Action Needed", "Follow Up Queue", "Applications", "Rejections", "Job Board Spam", "Important", "All Scanned"],
        )
        rows_by_key: Dict[str, Dict[str, Any]] = {}
        for _, _, row in parsed:
            key = row_processed_key(row)
            if key:
                rows_by_key[key] = row

        path = processed_log_path(project_dir, config)
        added = mark_rows_processed(
            path,
            list(rows_by_key.values()),
            run_meta={"command": "mark-report-processed", "report": str(Path(args.report))},
        )
        print(f"Marked {added} new message(s) as processed in: {path}")

    elif args.command == "processed-status":
        path = processed_log_path(project_dir, config)
        log = load_json_dict(path)
        print(f"Processed log: {path}")
        print(f"Processed messages: {len(log)}")

    elif args.command == "reset-processed":
        path = processed_log_path(project_dir, config)
        if path.exists():
            path.unlink()
            print(f"Deleted {path}")
        else:
            print(f"No processed log found at {path}")

    elif args.command == "send-approved":
        client = get_client(config, project_dir)
        process_approved_messages(client, Path(args.report), mode="send", config=config, yes_really=args.yes_really)
    elif args.command == "draft-approved":
        client = get_client(config, project_dir)
        process_approved_messages(client, Path(args.report), mode="draft", config=config, yes_really=False)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted.")
        sys.exit(130)
